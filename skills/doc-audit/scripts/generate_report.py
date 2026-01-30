#!/usr/bin/env python3
"""
ABOUTME: Generates HTML and Excel audit reports from audit manifest
ABOUTME: Includes statistics, issue details, and source tracing
"""

import argparse
import json
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

try:
    from jinja2 import Environment
except ImportError:
    print("Error: jinja2 not installed. Run: pip install jinja2", file=sys.stderr)
    sys.exit(1)

# Optional Excel support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def sanitize_excel_string(text: str) -> str:
    """
    Remove control characters that are illegal in Excel/XML.

    Excel uses XML internally, which only allows:
    #x9 (tab), #xA (LF), #xD (CR), and #x20-#xD7FF, #xE000-#xFFFD, #x10000-#x10FFFF

    Args:
        text: Text that may contain control characters

    Returns:
        Sanitized text safe for Excel
    """
    if not text or not isinstance(text, str):
        return text
    # Remove illegal control characters (0x00-0x08, 0x0B, 0x0C, 0x0E-0x1F)
    # Keep: \t (0x09), \n (0x0A), \r (0x0D)
    illegal_chars = ''.join(
        chr(c) for c in range(0x20)
        if c not in (0x09, 0x0A, 0x0D)
    )
    return text.translate(str.maketrans('', '', illegal_chars))


def load_manifest(file_path: str) -> tuple:
    """
    Load audit results and metadata from manifest JSONL file.

    Args:
        file_path: Path to manifest file

    Returns:
        Tuple of (metadata dict, list of audit result dictionaries)
    """
    metadata = {}
    results = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line:
                entry = json.loads(line)
                # Check if this is metadata
                if entry.get('type') == 'meta':
                    metadata = entry
                else:
                    results.append(entry)
    return metadata, results


def load_rules(file_path: str) -> dict:
    """
    Load rules from JSON file.

    Args:
        file_path: Path to rules JSON file

    Returns:
        Dictionary mapping rule_id to rule details
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        rules_data = json.load(f)
    
    rules_dict = {}
    for rule in rules_data.get('rules', []):
        rule_id = rule.get('id')
        if rule_id:
            rules_dict[rule_id] = {
                'id': rule_id,
                'description': rule.get('description', ''),
                'severity': rule.get('severity', 'medium'),
                'category': rule.get('category', 'other')
            }
    return rules_dict


def merge_rules(rule_files: list) -> dict:
    """
    Load and merge rules from multiple JSON files.
    Later files override earlier ones for duplicate rule IDs.
    
    Args:
        rule_files: List of paths to rules JSON files
        
    Returns:
        Dictionary mapping rule_id to rule details
    """
    merged = {}
    for file_path in rule_files:
        file_path_obj = Path(file_path)
        if file_path_obj.exists():
            rules = load_rules(file_path)
            merged.update(rules)  # Later files override
        else:
            print(f"Warning: Rules file not found: {file_path}", file=sys.stderr)
    return merged


def generate_report_data(manifest: list, rules_file_dict: dict = None) -> dict:
    """
    Generate report data from manifest.

    Args:
        manifest: List of audit result entries
        rules_file_dict: Optional dictionary of rules loaded from rules file

    Returns:
        Dictionary with report data
    """
    violations = []
    category_counts = Counter()
    rules = {}
    # Store unique source contents with source_id as key (uuid|uuid_end)
    source_contents = {}

    if rules_file_dict is None:
        rules_file_dict = {}

    for entry in manifest:
        if not entry.get('is_violation', False):
            continue

        # Handle multiple violations per block
        # Process violations array (new format)
        for v in entry.get('violations', []):
            category = v.get('category', 'other')
            rule_id = v.get('rule_id', '')

            uuid_start = entry.get('uuid', '')
            uuid_end = v.get('uuid_end', entry.get('uuid_end', ''))
            source_id = f"{uuid_start}|{uuid_end}"
            content = entry.get('p_content', '')

            # Store source content only once per unique source_id
            if source_id not in source_contents:
                source_contents[source_id] = content

            violations.append({
                'uuid': uuid_start,
                'uuid_end': uuid_end,  # Required for apply_audit_edits.py
                'source_id': source_id,  # Reference to source_contents
                'heading': entry.get('p_heading', ''),
                'category': category,
                'rule_id': rule_id,
                'violation_text': v.get('violation_text', ''),
                'violation_reason': v.get('violation_reason', ''),
                'fix_action': v.get('fix_action', 'manual'),
                'revised_text': v.get('revised_text', '')
            })

            category_counts[category] += 1

            # Collect unique rule information
            if rule_id and rule_id not in rules:
                # Try to get rule info from rules file first
                if rule_id in rules_file_dict:
                    rules[rule_id] = rules_file_dict[rule_id].copy()
                else:
                    # Fallback to data from manifest
                    rules[rule_id] = {
                        'id': rule_id,
                        'category': category,
                        'severity': v.get('severity', 'medium'),
                        'description': v.get('rule_description', '')
                    }

    return {
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'total_blocks': len(manifest),
        'violation_count': len(violations),
        'violations': violations,
        'category_counts': dict(category_counts),
        'max_category_count': max(category_counts.values()) if category_counts else 0,
        'rules': rules,
        'source_contents': source_contents  # Deduplicated source texts
    }


def render_report(data: dict, template_path: str, trusted_html: bool = False) -> str:
    """
    Render HTML report from data using Jinja2 template.

    Args:
        data: Report data dictionary
        template_path: Path to Jinja2 template file (required)
        trusted_html: If True, disable HTML escaping (use only for trusted inputs)

    Returns:
        Rendered HTML string
    """
    if not template_path:
        raise ValueError("Template path is required.")
    template_file = Path(template_path)
    if not template_file.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")
    template_str = template_file.read_text(encoding='utf-8')

    env = Environment(autoescape=not trusted_html)
    template = env.from_string(template_str)
    return template.render(**data)


def generate_excel_report(data: dict, output_path: str) -> None:
    """
    Generate Excel report from audit data.

    Args:
        data: Report data dictionary containing violations
        output_path: Path to save the Excel file

    Raises:
        ImportError: If openpyxl is not installed
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Report"

    # Define headers (Chinese)
    headers = [
        "序号",           # Index
        "错误类型",       # Error Type (category)
        "规则id",         # Rule ID
        "文本块标题",     # Text Block Title (heading)
        "违规原因",       # Violation Reason
        "操作建议",       # Operation Suggestion (fix_action)
        "违规原文",       # Violation Original Text
        "订正或建议"      # Correction (revised_text)
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Content alignment
    content_alignment = Alignment(vertical="top", wrap_text=True)

    # Write violation data
    for row_idx, violation in enumerate(data['violations'], 2):
        correction = violation.get('revised_text', '')

        row_data = [
            row_idx - 1,                              # 序号 (1-based index)
            violation.get('category', ''),            # 错误类型
            violation.get('rule_id', ''),             # 规则id
            violation.get('heading', ''),             # 文本块标题
            violation.get('violation_reason', ''),    # 违规原因
            violation.get('fix_action', ''),          # 操作建议
            violation.get('violation_text', ''),      # 违规原文
            correction                                # 订正或建议
        ]

        for col_idx, value in enumerate(row_data, 1):
            # Sanitize string values to remove illegal control characters
            safe_value = sanitize_excel_string(value) if isinstance(value, str) else value
            cell = ws.cell(row=row_idx, column=col_idx, value=safe_value)
            cell.alignment = content_alignment
            cell.border = thin_border

    # Auto-adjust column widths
    column_widths = [8, 15, 15, 25, 40, 15, 40, 40]  # Approximate widths
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Save workbook
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Generate HTML audit report from manifest"
    )
    parser.add_argument(
        "--manifest", "-m",
        type=str,
        required=True,
        help="Path to audit manifest JSONL file"
    )
    parser.add_argument(
        "--output", "-o",
        type=str,
        default="audit_report.html",
        help="Output HTML file path (default: audit_report.html)"
    )
    parser.add_argument(
        "--template", "-t",
        type=str,
        required=True,
        help="Path to Jinja2 HTML template (required)"
    )
    parser.add_argument(
        "--trusted-html",
        action="store_true",
        help="Render report without HTML escaping (only for trusted inputs)"
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="Also output report data as JSON"
    )
    parser.add_argument(
        "--rules", "-r",
        type=str,
        action='extend',
        nargs='+',
        default=[],
        help="Path to audit rules JSON file(s). Can be specified multiple times to merge rules."
    )
    parser.add_argument(
        "--excel",
        action="store_true",
        help="Also output report as Excel file (.xlsx)"
    )

    args = parser.parse_args()

    # Validate input
    manifest_path = Path(args.manifest)
    if not manifest_path.exists():
        print(f"Error: Manifest file not found: {args.manifest}", file=sys.stderr)
        sys.exit(1)

    template_path = Path(args.template)
    if not template_path.exists():
        print(f"Error: Template file not found: {args.template}", file=sys.stderr)
        sys.exit(1)

    # Load and merge rules if provided
    rules_dict = {}
    if args.rules:
        print(f"Loading rules from {len(args.rules)} file(s)...")
        rules_dict = merge_rules(args.rules)
        print(f"Loaded {len(rules_dict)} unique rules")

    # Load and process manifest
    print(f"Loading manifest: {args.manifest}")
    metadata, manifest = load_manifest(args.manifest)
    print(f"Loaded {len(manifest)} entries")
    
    # Display source file info if available
    if metadata:
        print(f"Source file: {metadata.get('source_file', 'Unknown')}")
        print(f"File hash: {metadata.get('source_hash', 'Unknown')[:20]}...")

    # Generate report data
    data = generate_report_data(manifest, rules_dict)
    
    # Add metadata to report data
    if metadata:
        data['source_file'] = metadata.get('source_file', 'Unknown')
        data['source_hash'] = metadata.get('source_hash', '')
        data['parsed_at'] = metadata.get('parsed_at', '')
        data['audited_at'] = metadata.get('audited_at', '')
    else:
        data['source_file'] = 'Unknown'
        data['source_hash'] = ''
        data['parsed_at'] = ''
        data['audited_at'] = ''
    
    print(f"Found {data['violation_count']} issues")

    # Render HTML
    html = render_report(data, args.template, trusted_html=args.trusted_html)

    # Save HTML
    output_path = Path(args.output)
    output_path.write_text(html, encoding='utf-8')
    print(f"HTML report saved to: {output_path}")

    # Optionally save JSON
    if args.json:
        json_path = output_path.with_suffix('.json')
        json_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding='utf-8')
        print(f"JSON data saved to: {json_path}")

    # Optionally save Excel
    if args.excel:
        if not EXCEL_AVAILABLE:
            print("Warning: openpyxl not installed. Skipping Excel output.", file=sys.stderr)
            print("Install with: pip install openpyxl", file=sys.stderr)
        else:
            excel_path = output_path.with_suffix('.xlsx')
            generate_excel_report(data, str(excel_path))
            print(f"Excel report saved to: {excel_path}")

    # Summary
    print("\n--- Summary ---")
    print(f"Total blocks: {data['total_blocks']}")
    print(f"Issues found: {data['violation_count']}")
    print(f"By category: {dict(data['category_counts'])}")


if __name__ == "__main__":
    main()
